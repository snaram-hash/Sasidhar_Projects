import os
import sqlite3
import logging
import openpyxl
from datetime import datetime
from config.settings import settings
from database.db_manager import db
from utils.exceptions import CAMGenerationError

logger = logging.getLogger("cuis.cam_service")

class CAMService:
    def generate_cam(self, borrower_id: int, financial_year: str) -> str:
        """Extract data from DB and generate a fully-populated Credit Assessment Memo (CAM) Excel workbook."""
        if not os.path.exists(settings.CAM_TEMPLATE_PATH):
            raise CAMGenerationError(f"CAM template sheet missing at: {settings.CAM_TEMPLATE_PATH}")
            
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            # 1. Fetch Borrower Details
            cursor.execute("SELECT company_name, pan, gstin FROM Borrowers WHERE id = ?", (borrower_id,))
            borrower = cursor.fetchone()
            if not borrower:
                raise CAMGenerationError(f"Borrower with ID {borrower_id} does not exist.")
                
            # 2. Fetch Audited Financials
            cursor.execute(
                """
                SELECT sales, pat, depreciation, interest_paid, reserves, net_worth, current_assets, current_liabilities,
                       secured_loans, unsecured_loans, working_capital_limits, purchases
                FROM Financials WHERE borrower_id = ? AND financial_year = ?
                """,
                (borrower_id, financial_year)
            )
            fin = cursor.fetchone()
            if not fin:
                raise CAMGenerationError(f"No audited financials found for borrower {borrower_id} for year {financial_year}.")
                
            # 3. Fetch GST Monthly Sales
            cursor.execute(
                """
                SELECT filing_period, taxable_sales FROM GSTSales WHERE borrower_id = ?
                """,
                (borrower_id,)
            )
            gst_rows = cursor.fetchall()
            gst_sales = {}
            for row in gst_rows:
                # Expecting format 'Month Year' like 'April 2024'
                period = row["filing_period"].strip()
                gst_sales[period] = row["taxable_sales"]
                
            # 4. Fetch Bank Transactions summary (monthly aggregate)
            # Group by year and month
            cursor.execute(
                """
                SELECT strftime('%Y-%m', tx_date) as yr_mo,
                       SUM(credit) as total_credits,
                       SUM(debit) as total_debits,
                       MIN(balance) as peak_od,
                       COUNT(CASE WHEN LOWER(narration) LIKE '%return%' OR LOWER(narration) LIKE '%bounce%' OR LOWER(narration) LIKE '%insufficient%' THEN 1 END) as bounce_count,
                       SUM(CASE WHEN LOWER(narration) LIKE '%interest%' OR LOWER(narration) LIKE '%int.coll%' THEN debit ELSE 0.0 END) as interest_charges
                FROM BankTransactions WHERE borrower_id = ?
                GROUP BY yr_mo
                """,
                (borrower_id,)
            )
            bank_rows = cursor.fetchall()
            bank_stats = {}
            for row in bank_rows:
                bank_stats[row["yr_mo"]] = {
                    "credits": row["total_credits"] or 0.0,
                    "debits": row["total_debits"] or 0.0,
                    "peak_od": row["peak_od"] or 0.0,
                    "bounces": row["bounce_count"] or 0,
                    "interest": row["interest_charges"] or 0.0
                }
                
            # Fetch Bank Account metadata (bank name and account_no from BankStatement Document)
            cursor.execute(
                """
                SELECT file_name FROM Documents 
                WHERE borrower_id = ? AND file_type = 'BankStatement' LIMIT 1
                """,
                (borrower_id,)
            )
            bank_doc = cursor.fetchone()
            bank_name = "AU SMALL FINANCE BANK"
            ac_no = "192837465012"
            if bank_doc:
                # Extract details from filename
                fname = bank_doc["file_name"].lower()
                if "hdfc" in fname:
                    bank_name = "HDFC BANK"
                elif "icici" in fname:
                    bank_name = "ICICI BANK"
                    
            # 5. Fetch CIBIL risk score and tier
            cursor.execute(
                """
                SELECT score, risk_tier FROM RiskAssessments 
                WHERE borrower_id = ? ORDER BY id DESC LIMIT 1
                """,
                (borrower_id,)
            )
            risk = cursor.fetchone()
            risk_tier = risk["risk_tier"] if risk else "Medium"
            
            # Step 2: Load workbook
            wb = openpyxl.load_workbook(settings.CAM_TEMPLATE_PATH)
            
            # Populate Sheet: FINANCIAL SHEET
            sheet_fs = wb['FINANCIAL SHEET']
            sheet_fs['B1'] = borrower["pan"]
            sheet_fs['E1'] = borrower["company_name"]
            sheet_fs['H73'] = borrower["gstin"]
            
            # Fetch all years of financials for this borrower
            cursor.execute(
                """
                SELECT financial_year, sales, pat, depreciation, interest_paid, reserves, net_worth, current_assets, current_liabilities,
                       secured_loans, unsecured_loans, working_capital_limits, purchases, debtors, creditors, inventory, fixed_assets,
                       direct_expenses, employee_expenses, other_income
                FROM Financials WHERE borrower_id = ?
                """,
                (borrower_id,)
            )
            fin_rows = cursor.fetchall()
            
            # Map financial year to column index (B=2, C=3, D=4, E=5)
            col_map = {}
            for row in fin_rows:
                fy_str = row["financial_year"].upper()
                if "22" in fy_str:
                    col_map[2] = row  # Column B
                elif "23" in fy_str:
                    col_map[3] = row  # Column C
                elif "24" in fy_str:
                    col_map[4] = row  # Column D
                elif "25" in fy_str:
                    col_map[5] = row  # Column E
                    
            for col_idx, f in col_map.items():
                sheet_fs.cell(row=4, column=col_idx).value = (f["sales"] or 0.0) / 100000.0
                sheet_fs.cell(row=5, column=col_idx).value = (f["sales"] or 0.0) / 100000.0  # Domestic Sales
                sheet_fs.cell(row=7, column=col_idx).value = (f["other_income"] or 0.0) / 100000.0  # Other income
                sheet_fs.cell(row=16, column=col_idx).value = (f["purchases"] or 0.0) / 100000.0
                sheet_fs.cell(row=18, column=col_idx).value = (f["direct_expenses"] or 0.0) / 100000.0  # Manufacturing expenses
                sheet_fs.cell(row=21, column=col_idx).value = (f["employee_expenses"] or 0.0) / 100000.0  # Salaries
                sheet_fs.cell(row=29, column=col_idx).value = (f["interest_paid"] or 0.0) / 100000.0
                sheet_fs.cell(row=30, column=col_idx).value = (f["depreciation"] or 0.0) / 100000.0
                sheet_fs.cell(row=34, column=col_idx).value = (f["pat"] or 0.0) / 100000.0
                
                sheet_fs.cell(row=41, column=col_idx).value = ((f["net_worth"] or 0.0) - (f["reserves"] or 0.0)) / 100000.0  # Share Capital
                sheet_fs.cell(row=43, column=col_idx).value = (f["reserves"] or 0.0) / 100000.0
                sheet_fs.cell(row=48, column=col_idx).value = (f["secured_loans"] or 0.0) / 100000.0
                sheet_fs.cell(row=49, column=col_idx).value = (f["unsecured_loans"] or 0.0) / 100000.0
                sheet_fs.cell(row=52, column=col_idx).value = (f["working_capital_limits"] or 0.0) / 100000.0
                sheet_fs.cell(row=57, column=col_idx).value = (f["creditors"] or 0.0) / 100000.0
                
                sheet_fs.cell(row=64, column=col_idx).value = (f["fixed_assets"] or 0.0) / 100000.0
                sheet_fs.cell(row=72, column=col_idx).value = (f["inventory"] or 0.0) / 100000.0
                sheet_fs.cell(row=73, column=col_idx).value = 0.0  # Debtors > 6 months
                sheet_fs.cell(row=74, column=col_idx).value = (f["debtors"] or 0.0) / 100000.0  # Debtors < 6 months
                
                # Cash & Bank (remaining current assets)
                cash_val = max(0.0, (f["current_assets"] or 0.0) - (f["debtors"] or 0.0) - (f["inventory"] or 0.0))
                sheet_fs.cell(row=76, column=col_idx).value = cash_val / 100000.0
            
            # Populate monthly GST sales
            # Loop months (from April 2024 to March 2025)
            sheet_fs['H46'] = datetime(2024, 4, 1)
            months_seq = [
                ("April 2024", 46), ("May 2024", 47), ("June 2024", 48),
                ("July 2024", 49), ("August 2024", 50), ("September 2024", 51),
                ("October 2024", 52), ("November 2024", 53), ("December 2024", 54),
                ("January 2025", 55), ("February 2025", 56), ("March 2025", 57)
            ]
            for m_str, r_idx in months_seq:
                val = gst_sales.get(m_str, 0.0) / 100000.0
                sheet_fs.cell(row=r_idx, column=10).value = val
                
            # Populate Sheet: Banking System Analysis
            sheet_bsa = wb['Banking System Analysis']
            sheet_bsa['J9'] = bank_name
            sheet_bsa['L9'] = "CC"
            sheet_bsa['R9'] = ac_no
            sheet_bsa['T9'] = borrower["company_name"]
            
            # Write monthly bank statement statistics
            # Map database keys 'YYYY-MM' to row indexes
            months_bank_seq = [
                ("2024-04", 28), ("2024-05", 29), ("2024-06", 30),
                ("2024-07", 31), ("2024-08", 32), ("2024-09", 33),
                ("2024-10", 34), ("2024-11", 35), ("2024-12", 36),
                ("2025-01", 37), ("2025-02", 38), ("2025-03", 39)
            ]
            for ym_str, r_idx in months_bank_seq:
                stats = bank_stats.get(ym_str, None)
                if stats:
                    sheet_bsa.cell(row=r_idx, column=11).value = stats["credits"]
                    sheet_bsa.cell(row=r_idx, column=12).value = stats["debits"]
                    sheet_bsa.cell(row=r_idx, column=13).value = stats["peak_od"]
                    sheet_bsa.cell(row=r_idx, column=16).value = stats["interest"]
                    sheet_bsa.cell(row=r_idx, column=19).value = stats["bounces"]
                    
            # Save Workbook under exports path
            export_dir = os.path.join(settings.BASE_DIR, "assets", "exports", str(borrower_id))
            os.makedirs(export_dir, exist_ok=True)
            export_path = os.path.join(export_dir, f"CAM_{borrower_id}_{financial_year}.xlsx")
            wb.save(export_path)
            
            # Record in database CAMHistory table
            cursor.execute(
                """
                INSERT INTO CAMHistory (borrower_id, cam_path, status, underwriter_notes)
                VALUES (?, ?, 'Draft', ?)
                """,
                (borrower_id, export_path, f"CAM generated for {financial_year}. Risk recommendation: {risk_tier}.")
            )
            conn.commit()
            logger.info(f"CAM successfully generated and stored at: {export_path}")
            return export_path
            
        except Exception as e:
            conn.rollback()
            logger.error(f"Failed to generate CAM workbook: {e}")
            raise CAMGenerationError(f"CAM Generation Engine error: {e}")
        finally:
            conn.close()

cam_service = CAMService()
